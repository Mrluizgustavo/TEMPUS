import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment
from .processador import ResultadoJornada


# Amarelo de atenção — visível sem ser agressivo
_FILL_REVISAR = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_FILL_HEADER  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_FONT_HEADER  = Font(bold=True)


class ExcelReporterRevisao:
    """
    Gera o Excel intermediário de revisão humana.
    Contém apenas a segmentação das jornadas — sem status, duração ou intervalo.
    O usuário corrige os agrupamentos e depois confirma na Etapa 2.

    Coluna REVISAR:
        "⚠ Múltiplas jornadas no dia" — quando o mesmo funcionário tem mais de
        uma jornada registrada na mesma data, indicando possível erro de
        segmentação. A linha inteira é destacada em amarelo para facilitar
        a localização visual durante a revisão.
    """

    NOME_ARQUIVO  = "Revisao_Jornadas.xlsx"
    NOME_ABA      = "REVISÃO"
    CAMINHO_PASTA = "data/output"

    @classmethod
    def caminho_revisao(cls) -> str:
        return os.path.abspath(os.path.join(cls.CAMINHO_PASTA, cls.NOME_ARQUIVO))

    def gerar_excel_revisao(self, resultados: list[ResultadoJornada]) -> str:
        if not resultados:
            print("⚠️ Nenhum dado para gerar revisão.")
            return ""

        max_batidas = max((len(r.batidas) for r in resultados), default=0)

        linhas = []
        for r in resultados:
            linha = {
                "CHAPA":   r.chapa,
                "LOJA":    r.loja,
                "NOME":    r.nome,
                "IDADE":   r.idade,
                "DATA":    r.data_inicio_str,
                "REVISAR": "⚠ Múltiplas jornadas no dia" if r.multipla_jornada_no_dia else "",
            }
            for i, horario in enumerate(r.batidas):
                # Remove prefixo de dia seguinte "(DD) " para deixar só "HH:MM"
                horario_limpo = horario
                if horario.startswith("("):
                    horario_limpo = horario[6:].strip()
                linha[f"BAT{i + 1}"] = horario_limpo

            linhas.append(linha)

        df = pd.DataFrame(linhas)

        # Garante todas as colunas de batida mesmo que algumas jornadas tenham menos
        colunas_fixas   = ["CHAPA", "LOJA", "NOME", "IDADE", "DATA", "REVISAR"]
        colunas_batidas = [f"BAT{i + 1}" for i in range(max_batidas)]
        df = df.reindex(columns=colunas_fixas + colunas_batidas).fillna("")

        os.makedirs(self.CAMINHO_PASTA, exist_ok=True)
        caminho_completo = os.path.join(self.CAMINHO_PASTA, self.NOME_ARQUIVO)

        try:
            with pd.ExcelWriter(caminho_completo, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=self.NOME_ABA, index=False)
                ws = writer.sheets[self.NOME_ABA]

                # Congela cabeçalho
                ws.freeze_panes = "A2"

                # Formata cabeçalho
                for cell in ws[1]:
                    cell.font      = _FONT_HEADER
                    cell.fill      = _FILL_HEADER
                    cell.alignment = Alignment(horizontal="center")

                # Índice da coluna REVISAR no Excel (1-based)
                idx_col_revisar = colunas_fixas.index("REVISAR") + 1

                # Pinta linhas sinalizadas (linha 2 em diante — linha 1 é cabeçalho)
                total_colunas = len(colunas_fixas) + len(colunas_batidas)
                for row_idx, resultado in enumerate(resultados, start=2):
                    if resultado.multipla_jornada_no_dia:
                        for col_idx in range(1, total_colunas + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = _FILL_REVISAR

                # Ajusta largura das colunas automaticamente
                for col in ws.columns:
                    max_len = max(
                        (len(str(cell.value)) for cell in col if cell.value), default=0
                    )
                    ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 10)

            print(f"✅ Excel de revisão gerado: {caminho_completo}")
            return caminho_completo

        except PermissionError:
            raise PermissionError(
                f"O arquivo '{self.NOME_ARQUIVO}' está aberto. Feche-o e tente novamente."
            )